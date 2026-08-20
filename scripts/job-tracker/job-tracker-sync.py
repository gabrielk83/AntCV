#!/usr/bin/env python3
# JOB-TRACKER-001 — local sync client for the AntCV job-search workbook.
# Syncs a canonical job_tracker_doc.json <-> the D1-backed /api/job-tracker
# endpoint on the access-relay. D1 is the source of truth; the Excel workbook
# is rendered FROM the doc (build script, doc-driven via JOB_DOC).
#
# This tool is GENERIC (no personal data). The doc + the Excel build script
# live locally (not committed) per the repo's "no candidate data" rule.
#
# Auth: the relay verifies a JWT (the same token the PWA holds after sign-in).
#   Provide it via  env ANTCV_TOKEN  or a file (default ~/.antcv/token).
#   Get it from the PWA: DevTools > Application > Local Storage > your token,
#   or however you export your session key.
#
# Commands:
#   pull            GET the cloud doc, 3-way MERGE it into the local doc (local-only entries are
#                   never dropped; --force takes the cloud wholesale) + write snapshot;
#                   --render also rebuilds the .xlsx
#   push            PUT the local doc with the snapshot's base_rev; on 409 do a row-level 3-way merge and retry
#   import-xlsx      read the (edited) Weekly Tracker sheet back into the local doc, then run `push`
#   status           show local rev vs cloud rev and whether they diverge
#
# Config via env (all optional):
#   ANTCV_RELAY   default https://antcv-access-relay.karp-gabriel-a.workers.dev
#   ANTCV_TOKEN   JWT bearer token (else read ANTCV_TOKEN_FILE)
#   ANTCV_TOKEN_FILE  default ~/.antcv/token
#   JOB_DOC       path to job_tracker_doc.json (canonical)
#   JOB_BUILD     path to the local doc-driven Excel build script (for --render)
#   JOB_XLSX      path to the .xlsx (for import-xlsx)
import os, sys, json, hashlib, subprocess, urllib.request, urllib.error

RELAY = os.environ.get("ANTCV_RELAY", "https://antcv-access-relay.karp-gabriel-a.workers.dev").rstrip("/")
DOC   = os.environ.get("JOB_DOC",  os.path.expanduser("~/job_tracker_doc.json"))
BUILD = os.environ.get("JOB_BUILD")
XLSX  = os.environ.get("JOB_XLSX")
SNAP  = DOC + ".sync"          # snapshot: {"rev": int, "doc": {...}}
ENDPOINT = RELAY + "/api/job-tracker"

def _token_file():
    return os.environ.get("ANTCV_TOKEN_FILE", os.path.expanduser("~/.antcv/token"))

def _token():
    t = os.environ.get("ANTCV_TOKEN")
    if t: return t.strip()
    p = _token_file()
    if os.path.exists(p):
        return open(p, "r", encoding="utf-8").read().strip()
    sys.exit("No token. Set ANTCV_TOKEN or put it in " + p)

# Persist a rotated token so ~/.antcv/token renews itself and never expires
# while the CLI/nightly keeps running (the relay sends X-Auth-Refresh once the
# token is >1 day old). Only writes the file form (an env-var token can't be
# updated from here).
def _save_token(t):
    if not t or os.environ.get("ANTCV_TOKEN"): return
    try:
        p = _token_file(); os.makedirs(os.path.dirname(p), exist_ok=True)
        open(p, "w", encoding="utf-8").write(t.strip())
    except Exception: pass

def _req(method, body=None):
    data = json.dumps(body).encode() if body is not None else None
    r = urllib.request.Request(ENDPOINT, data=data, method=method,
        headers={"Authorization": "Bearer " + _token(),
                 "Content-Type": "application/json",
                 "Origin": "https://antcv.pages.dev",
                 # Cloudflare's edge blocks the default Python-urllib UA (403);
                 # present a browser-like UA so the request reaches the worker.
                 "User-Agent": "Mozilla/5.0 (AntCV job-tracker-sync)"})
    try:
        with urllib.request.urlopen(r, timeout=30) as resp:
            _save_token(resp.headers.get("X-Auth-Refresh"))
            return resp.status, json.loads(resp.read().decode() or "{}")
    except urllib.error.HTTPError as e:
        try: payload = json.loads(e.read().decode() or "{}")
        except Exception: payload = {}
        return e.code, payload

def _load(p, default=None):
    return json.load(open(p, "r", encoding="utf-8")) if os.path.exists(p) else default

def _save(p, obj):
    d = os.path.dirname(os.path.abspath(p))
    try:
        os.makedirs(d, exist_ok=True)
    except OSError as e:
        # JOBTRACKER-DRIVE-UNMOUNTED-001: a raw WinError 3 here means the drive
        # root itself doesn't exist (e.g. a mapped/virtual drive like Google
        # Drive for Desktop isn't mounted in this session) — not a normal
        # missing-subfolder case makedirs would otherwise handle. Surface that
        # plainly instead of an opaque traceback so an unattended run's log
        # says what's actually wrong.
        sys.exit(f"cannot write {p}: {d} is unreachable ({e}). "
                  f"Is the drive mounted (e.g. Google Drive for Desktop running "
                  f"and signed in in this session)?")
    json.dump(obj, open(p, "w", encoding="utf-8"), ensure_ascii=False, indent=1)

def _rows_by_id(doc):
    # row id = urlkey (index 11 in the tuple/list); fall back to company+role.
    out = {}
    for row in (doc.get("rows") or []):
        rid = row[11] if len(row) > 11 and row[11] else (str(row[1]) + "|" + str(row[2]))
        out[rid] = row
    return out

def cmd_status():
    code, remote = _req("GET")
    snap = _load(SNAP, {"rev": 0})
    local = _load(DOC)
    print(f"cloud rev: {remote.get('rev') if code==200 else 'ERR '+str(code)}")
    print(f"snapshot rev: {snap.get('rev', 0)}")
    print(f"local doc: {'present' if local else 'MISSING'} ({DOC})")
    if code == 200 and local:
        diverged = remote.get("rev") != snap.get("rev")
        localdirty = _load(SNAP, {}).get("doc") != local
        print(f"cloud advanced since last sync: {diverged}")
        print(f"local edited since last sync:  {localdirty}")

# Tier band hex -> a short human label (mirrors the JobTracker island TIERS map).
_TIER_LABEL = {"DDEBF7": "T1 strong fit", "E2EFDA": "T2 transferable",
               "FCE4D6": "T3 weak / pivot", "FFF2CC": "In progress", "D9D9D9": "Archive"}

def _write_proposed_tab(xlsx_path, doc):
    """Add/refresh a 'Proposed / Inbox' worksheet = a focused triage view of the
    auto-discovered leads (rows whose GROUP cell == 'Proposed', written by
    discover-positions.py). This is a READ-ONLY view: import-xlsx only reads the
    'Weekly Tracker' sheet, so this tab never affects the doc round-trip. It runs
    AFTER build_workbook.py on every --render, so build's full-file rewrite can
    never leave a stale copy. Returns the number of proposed leads written."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    SHEET = "Proposed Inbox"
    urls = doc.get("urls") or {}
    seen_by_uk = {}
    for v in (doc.get("discovered") or {}).values():
        if isinstance(v, dict) and v.get("uk"):
            seen_by_uk[v["uk"]] = v.get("first_seen", "")
    proposed = [r for r in (doc.get("rows") or [])
                if len(r) > 5 and str(r[5]).strip().lower() == "proposed"]
    wb = load_workbook(xlsx_path)
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET, 1)  # right after Weekly Tracker
    hdr_fill = PatternFill("solid", fgColor="1F3A5F"); hdr_font = Font(bold=True, color="FFFFFF")
    ws.cell(2, 2, "PROPOSED — auto-discovered leads (Sun & Tue night). Review, then in the "
                  "AntCV app: pin/generate the good ones, or reject with a reason. Rebuilds each sync.")
    cols = ["#", "Company", "Role", "Location", "Tier", "Fit", "Posting",
            "Why (auto-discovered)", "Status", "First seen"]
    for ci, h in enumerate(cols, start=1):
        c = ws.cell(4, ci, h); c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(horizontal="left", vertical="center")
    ri = 5
    for r in sorted(proposed, key=lambda x: (str(x[12] if len(x) > 12 else ""), x[0])):
        uk = r[11] if len(r) > 11 else ""
        band = str((r[12] if len(r) > 12 else "") or "E2EFDA").upper()
        vals = [r[0], r[1], r[2], r[3] if len(r) > 3 else "", _TIER_LABEL.get(band, band),
                r[6] if len(r) > 6 else "", "", r[10] if len(r) > 10 else "",
                r[8] if len(r) > 8 else "", seen_by_uk.get(uk, "")]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(ri, ci, v)
            c.fill = PatternFill("solid", fgColor=band)
            c.alignment = Alignment(vertical="top", wrap_text=(ci in (2, 3, 8)))
        url = urls.get(uk, "")
        if url:
            pc = ws.cell(ri, 7, "Open posting"); pc.hyperlink = url
            pc.font = Font(color="1155CC", underline="single")
        ri += 1
    if not proposed:
        ws.cell(6, 2, "No proposed leads right now — the discovery task adds them Sun & Tue night.")
    for ci, w in enumerate([5, 22, 30, 16, 16, 24, 14, 42, 22, 12], start=1):
        ws.column_dimensions[chr(64 + ci)].width = w
    ws.freeze_panes = "A5"
    wb.save(xlsx_path)
    return len(proposed)

def cmd_pull(render=False, force=False):
    code, remote = _req("GET")
    if code != 200: sys.exit(f"pull failed: {code} {remote}")
    doc = remote.get("doc")
    if doc is None:
        print("cloud is empty — nothing to pull. (Run push to seed it.)")
        return
    # JOBTRACKER-PULL-CLOBBERS-LOCAL-SUPPORT-001 (2026-08-20): `pull` used to
    # _save(DOC, doc) the cloud doc straight over the local file. `push` has a
    # careful 3-way merge; `pull` had none, so anything local the cloud had not
    # seen was silently deleted. Measured on 2026-08-20: the local doc held 34
    # `support` blocks of role intel the cloud lacked - one nightly `pull
    # --render` would have destroyed all of them. Pull now runs the SAME 3-way
    # merge push does. `--force` restores the old take-the-cloud-wholesale
    # behaviour for when that is genuinely what you want.
    local = _load(DOC)
    forced = force or not local
    if forced:
        merged, conflicts = doc, []
    else:
        merged, conflicts = _merge(_load(SNAP, {}).get("doc"), local, doc)
    _save(DOC, merged)
    # The SNAPSHOT stays a faithful record of what the CLOUD holds at this rev,
    # so a later push sends the right base_rev and diffs against the right base.
    _save(SNAP, {"rev": remote.get("rev", 0), "doc": doc})
    print(f"pulled rev {remote.get('rev')} -> {DOC}" + (" (--force: local overwritten)" if forced else ""))
    if not forced:
        kept = sum(1 for f in UK_DICT_FIELDS
                   for k in ((local or {}).get(f) or {})
                   if k not in ((doc or {}).get(f) or {}))
        extra_rows = len(merged.get("rows") or []) - len(doc.get("rows") or [])
        if kept or extra_rows > 0:
            print(f"merged: kept {kept} local-only entr{'y' if kept == 1 else 'ies'}"
                  + (f" + {extra_rows} local-only row(s)" if extra_rows > 0 else "")
                  + " the cloud does not have - run `push` to upload them")
        if conflicts:
            print(f"conflicts (local kept): {', '.join(conflicts[:12])}"
                  + (f" +{len(conflicts) - 12} more" if len(conflicts) > 12 else ""))
    if render:
        if not BUILD: sys.exit("--render needs JOB_BUILD (path to the doc-driven Excel build script)")
        env = dict(os.environ, JOB_DOC=DOC)
        subprocess.run([sys.executable, BUILD], check=True, env=env)
        print("rebuilt .xlsx from pulled doc")
        # PROPOSED-INBOX-TAB-001: append the focused discovery-inbox view. Runs
        # after build (which rewrites the whole file), guarded so it never breaks
        # a render. Needs JOB_XLSX to locate the built workbook.
        if XLSX:
            try:
                n = _write_proposed_tab(XLSX, doc)
                print(f"added 'Proposed Inbox' tab ({n} lead{'s' if n != 1 else ''})")
            except Exception as e:
                # JOBTRACKER-PROPOSED-TAB-PATH-QUOTE-001: the old [:90] slice cut a
                # FileNotFoundError mid-path, making a wrong JOB_XLSX value (real
                # cause) misread as a path-parsing bug. Print it whole.
                print(f"proposed-tab skipped ({type(e).__name__}: {e})")
        else:
            print("proposed-tab skipped (set JOB_XLSX to enable)")

# Every uk-keyed dict in the doc besides `rows`. These carry the JD text, the
# role intel, the owner's typed signals and the generation state - i.e. most of
# the doc's value - and until 2026-08-20 the sync merged NONE of them
# (JOBTRACKER-PULL-CLOBBERS-LOCAL-SUPPORT-001).
UK_DICT_FIELDS = ("urls", "jd", "support", "signals", "notes", "webintel", "gen",
                  "queue", "brandfit", "brand", "artifacts", "sigfiles", "pin",
                  "park", "discovered")

def _merge_uk_dict(base, local, remote, field):
    """Per-KEY 3-way for one uk-keyed dict. Union of both sides - a key present
    on one side only is always kept, never deleted, because neither `pull` nor a
    409 `push` can tell "the other side removed it" from "the other side has not
    seen it yet". Where both sides changed the same key, LOCAL wins and the key
    is reported. Returns (merged, conflict_labels)."""
    b = (base or {}).get(field) or {}
    l = (local or {}).get(field) or {}
    r = (remote or {}).get(field) or {}
    out, conflicts = dict(r), []
    for k, lv in l.items():
        if k not in r:
            out[k] = lv                                     # local-only -> keep
            continue
        rv, bv = r[k], b.get(k)
        if lv == rv: continue
        lc, rc = (lv != bv), (rv != bv)
        if lc and not rc: out[k] = lv
        elif rc and not lc: out[k] = rv
        else: out[k] = lv; conflicts.append(f"{field}[{k}](local-wins)")
    return out, conflicts

def _merge(base, local, remote):
    # Row-level 3-way by id. For each id: if only one side changed vs base, take
    # that side; if both changed the same row differently, LOCAL wins and the id
    # is reported as a conflict (never silently dropped).
    b, l, r = _rows_by_id(base or {}), _rows_by_id(local), _rows_by_id(remote)
    ids = list(dict.fromkeys(list(r.keys()) + list(l.keys())))
    merged, conflicts = [], []
    for i in ids:
        bl, ll, rr = b.get(i), l.get(i), r.get(i)
        if ll is None: merged.append(rr); continue          # deleted locally? keep remote (safe)
        if rr is None: merged.append(ll); continue          # new local row
        if ll == rr: merged.append(ll); continue
        lc, rc = (ll != bl), (rr != bl)
        if lc and not rc: merged.append(ll)
        elif rc and not lc: merged.append(rr)
        else: merged.append(ll); conflicts.append(i)        # both changed -> local wins, flag
    out = dict(local)
    out["rows"] = merged
    # envelope: last-writer (local) wins whole-block; report if it differs from remote
    if remote.get("envelope") != local.get("envelope"): conflicts.append("envelope(local-wins)")
    for _f in UK_DICT_FIELDS:
        if _f in local or _f in remote or _f in (base or {}):
            merged_f, cf = _merge_uk_dict(base, local, remote, _f)
            out[_f] = merged_f
            conflicts.extend(cf)
    return out, conflicts

def cmd_push():
    local = _load(DOC)
    if not local: sys.exit(f"no local doc at {DOC}")
    snap = _load(SNAP, {"rev": 0, "doc": None})
    code, resp = _req("PUT", {"doc": local, "base_rev": snap.get("rev", 0)})
    if code == 200:
        _save(SNAP, {"rev": resp["rev"], "doc": local})
        print(f"pushed -> rev {resp['rev']}")
        return
    if code == 409:
        remote = resp.get("doc")
        print(f"conflict: cloud is at rev {resp.get('rev')} (snapshot {snap.get('rev')}). Merging…")
        merged, conflicts = _merge(snap.get("doc"), local, remote)
        code2, resp2 = _req("PUT", {"doc": merged, "base_rev": resp.get("rev")})
        if code2 != 200: sys.exit(f"merge push failed: {code2} {resp2}")
        _save(DOC, merged); _save(SNAP, {"rev": resp2["rev"], "doc": merged})
        print(f"merged & pushed -> rev {resp2['rev']}")
        if conflicts: print("CONFLICTS (local kept, review these ids):", ", ".join(conflicts))
        return
    sys.exit(f"push failed: {code} {resp}")

def cmd_import_xlsx():
    # Read the edited Weekly Tracker sheet back into the local doc, preserving
    # each row's url-key (col 12) + band colour (from the row fill) by matching
    # on rank -> then company+role. Then push.
    from openpyxl import load_workbook
    if not XLSX: sys.exit("import-xlsx needs JOB_XLSX (path to the .xlsx)")
    doc = _load(DOC) or {"version": 1, "rows": [], "envelope": [], "urls": {}}
    old = _rows_by_id(doc)
    by_rank = {row[0]: row for row in doc.get("rows", [])}
    wb = load_workbook(XLSX)
    ws = wb["Weekly Tracker"]
    # header row is 4; data starts row 5. Columns: 2=Rank..12=Flag (1-based).
    new_rows = []
    for r in range(5, ws.max_row + 1):
        rank = ws.cell(r, 2).value
        comp = ws.cell(r, 3).value
        if rank is None and not comp: continue
        vals = [ws.cell(r, c).value for c in range(2, 13)]  # rank..flag (11 fields)
        prev = by_rank.get(rank) or old.get(str(rank) + "|" + str(comp))
        uk = prev[11] if prev and len(prev) > 11 else (str(comp or "row").lower().split()[0])
        band = None
        f = ws.cell(r, 2).fill
        if f and f.fgColor and f.fgColor.rgb and str(f.fgColor.rgb) not in ("00000000",):
            band = str(f.fgColor.rgb)[-6:]
        if not band and prev and len(prev) > 12: band = prev[12]
        new_rows.append(list(vals) + [uk, band or "E2EFDA"])
    doc["rows"] = new_rows
    _save(DOC, doc)
    print(f"imported {len(new_rows)} tracker rows from .xlsx -> {DOC}")
    cmd_push()

def main():
    cmd = sys.argv[1] if len(sys.argv) > 1 else "status"
    if cmd == "pull": cmd_pull(render=("--render" in sys.argv), force=("--force" in sys.argv))
    elif cmd == "push": cmd_push()
    elif cmd == "import-xlsx": cmd_import_xlsx()
    elif cmd == "status": cmd_status()
    else: sys.exit("usage: job-tracker-sync.py [pull [--render] [--force] | push | import-xlsx | status]")

if __name__ == "__main__":
    main()
